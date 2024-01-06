"""
intermediate_assignment_63

Write a Python Program to write data to excel cell using python other than xlsxwriter library 

"""

#Create Excel Workbook and Add Worksheets using Python
from openpyxl import Workbook

wb = Workbook()
wb.create_sheet("Sheet_one")
wb.create_sheet("Sheet_two")

wb.sheetnames
['Sheet', 'Sheet_one', 'Sheet_two']

# Write Some Data Into Excel
ws1 = wb['Sheet_one']

ws1['B2'] = 'Fruits'
ws1['C2'] = 'Sales'
ws1['E4'] = 'Books'


#Saving Excel File
wb.save('book_eg.xlsx')

